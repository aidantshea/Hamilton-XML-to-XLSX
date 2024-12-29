"""
This script supports application.py. It contains the scrape and convert_XML_to_XLSX methods for import.
"""

from PySide6.QtWidgets import QMessageBox
import xml.etree.ElementTree as ET

# this method takes an xml file path, scrapes and formats specified data, and returns it as a 2D array
def scrape(xml_path):
    
    # reading given xml and saving text under SpecID and Well tags to respective lists
    tree = ET.parse(xml_path)
    root = tree.getroot()

    specimens = [elem.text for elem in root.iter('SampleID')]
    wells =[elem.text for elem in root.iter('SamplePos')]
    
    # formatting specimen data into a 2d array
    specimen_data = []
    for i in range(len(specimens)):
        specimen_data.append([specimens[i], wells[i]])
    
    # generating a message box for sanity check by user
    message = QMessageBox(); message.setWindowTitle("Sanity Check!")
    toprint = []
    if len(specimen_data) < 20:
        for i in range(len(specimen_data)):
            toprint.append(specimen_data[i][0] + " - " + specimen_data[i][1])
    else:
        for i in range(10):
            toprint.append(specimen_data[i][0] + " - " + specimen_data[i][1])
        toprint.append("\n")
        for i in range(10):
            toprint.append(specimen_data[len(specimen_data) - 10 + i][0] + " - " + specimen_data[len(specimen_data) - 10 + i][1])
    message.setText('\n'.join(toprint))
    message.exec()
    
    return specimen_data

from openpyxl import load_workbook
import os

# this method uses the soup to generate a plate map in XLSX format and saves to the specified location on the disc
def convert_XML_to_XLSX(outfile_path, specimen_data):
    
    # loading in template workbook, which should be saved in this script's directory
    basedir = os.path.dirname(__file__)
    workbook = load_workbook(os.path.join(basedir, "template.xlsx"), data_only=True)
    sheet = workbook.active
    
    # writing the specimen IDs into the template, then saving as a .xlsx
    for i in range(len(specimen_data)):
        for k in range(184):
            if specimen_data[i][1] == sheet["A" + str(k+9)].value:
                sheet.cell(row=k+9, column=2, value = specimen_data[i][0])
   
    workbook.save(outfile_path)